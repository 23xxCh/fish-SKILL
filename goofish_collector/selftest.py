from __future__ import annotations

import tempfile
import traceback
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .browser import GoofishBrowserSession
from .exporter import export_workbook
from .models import CrawlConfig, SearchFilters
from .parser import parse_card


def run_self_test() -> int:
    error_log = Path(tempfile.gettempdir()) / "XianyuLinkCollector-selftest.txt"
    try:
        with tempfile.TemporaryDirectory(prefix="goofish-selftest-") as temporary:
            root = Path(temporary)
            with GoofishBrowserSession(root / "profile", headless=True) as session:
                session.page.set_content(
                    """
                    <html><body>
                    <input placeholder="¥"><input placeholder="¥">
                    <button>确定</button>
                    <button>新发布</button><button>3天内</button>
                    <button>区域</button><button>广东</button>
                    <label><input type="checkbox">个人闲置</label>
                    <label><input type="checkbox">验货宝</label>
                    <label><input type="checkbox">包邮</label>
                    <label><input type="checkbox">全新</label>
                    <article>
                    <a href="https://www.goofish.com/item?id=123456789">自检商品</a>
                    <div>95新 ¥9 广东</div>
                    </article></body></html>
                    """
                )
                filters = SearchFilters(
                    min_price=100,
                    max_price=800,
                    region="广东",
                    published_within="3天内",
                    personal_only=True,
                    inspection_only=True,
                    free_shipping=True,
                    brand_new=True,
                )
                session.apply_filters(filters)
                payload = session.extract_cards()[0]
            record = parse_card(
                payload,
                keyword="自检商品",
                page=1,
                captured_at="2026-08-04 10:00:00",
            )
            if record is None:
                raise RuntimeError("自检卡片未解析为商品记录")
            config = CrawlConfig("自检商品", 1, 0, root, filters=filters)
            output = export_workbook(
                config=config,
                records=[record],
                raw_records=1,
                searched_pages=1,
                stop_reason="自检完成",
                started_at=datetime(2026, 8, 4, 10, 0, 0),
                finished_at=datetime(2026, 8, 4, 10, 0, 1),
            )
            workbook = load_workbook(output, read_only=False)
            link_cell = workbook["全部链接"]["P5"]
            valid = (
                workbook.sheetnames == ["汇总说明", "全部链接"]
                and workbook["汇总说明"]["B8"].value == "¥100.00 - ¥800.00"
                and workbook["汇总说明"]["B9"].value == "广东"
                and workbook["汇总说明"]["B10"].value
                == "3天内｜个人闲置、验货宝、包邮、全新"
                and link_cell.value == record.url
                and link_cell.hyperlink is not None
                and link_cell.hyperlink.target == record.url
            )
            workbook.close()
            if not valid:
                raise RuntimeError("自检 Excel 的工作表或超链接不正确")
        error_log.unlink(missing_ok=True)
        return 0
    except Exception:
        error_log.write_text(traceback.format_exc(), encoding="utf-8")
        return 1
