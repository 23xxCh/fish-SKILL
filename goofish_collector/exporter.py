from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import CrawlConfig, ProductRecord
from .monitor_models import MonitorTaskConfig


HEADERS = [
    "序号",
    "关键词",
    "商品ID",
    "标题",
    "价格",
    "原价",
    "地区",
    "成色",
    "想要人数",
    "卖家标签",
    "发布/降价动态",
    "累计降价",
    "首次出现页",
    "出现次数",
    "出现页码",
    "商品链接",
    "抓取时间",
    "原始商品文本",
]


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .")
    return (cleaned or "闲鱼商品")[:60]


def _unique_output_path(directory: Path, keyword: str, finished_at: datetime) -> Path:
    base = f"闲鱼_{_safe_filename(keyword)}_商品链接_{finished_at:%Y%m%d_%H%M%S}"
    candidate = directory / f"{base}.xlsx"
    suffix = 2
    while candidate.exists():
        candidate = directory / f"{base}_{suffix}.xlsx"
        suffix += 1
    return candidate


def _as_datetime(value: str) -> datetime | str:
    try:
        return datetime.fromisoformat(value)
    except (TypeError, ValueError):
        return value


def _style_summary(sheet, *, config: CrawlConfig, row_count: int) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "闲鱼商品链接采集汇总"
    sheet["A1"].font = Font(name="微软雅黑", size=20, bold=True, color="1F2937")
    sheet["A1"].fill = PatternFill("solid", fgColor="F8D447")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 38

    section_fill = PatternFill("solid", fgColor="2B3745")
    section_font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    sheet.merge_cells("A3:B3")
    sheet["A3"] = "任务信息"
    sheet["A3"].fill = section_fill
    sheet["A3"].font = section_font
    sheet.merge_cells("D3:E3")
    sheet["D3"] = "数据结果"
    sheet["D3"].fill = section_fill
    sheet["D3"].font = section_font

    labels = ["关键词", "最大页数", "最大商品数", "输出目录", "价格区间", "地区", "发布时间/商品条件"]
    for row, label in enumerate(labels, start=4):
        sheet.cell(row, 1, label)
    sheet["B4"] = config.keyword
    sheet["B5"] = config.max_pages
    sheet["B6"] = "不限" if config.max_items == 0 else config.max_items
    sheet["B7"] = str(config.output_dir)
    sheet["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.row_dimensions[7].height = 44
    sheet["B8"] = config.filters.price_label()
    sheet["B9"] = config.filters.region or "全国"
    sheet["B10"] = config.filters.other_label()
    sheet["B10"].alignment = Alignment(wrap_text=True, vertical="center")

    result_labels = ["原始展示记录", "唯一商品链接", "搜索页数", "停止原因", "开始时间", "完成时间"]
    for row, label in enumerate(result_labels, start=4):
        sheet.cell(row, 4, label)
    sheet["E5"] = row_count

    note_row = 12
    sheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    sheet.cell(note_row, 1, "说明")
    sheet.cell(note_row, 1).fill = section_fill
    sheet.cell(note_row, 1).font = section_font
    sheet.merge_cells(start_row=13, start_column=1, end_row=16, end_column=6)
    sheet["A13"] = (
        "1. “全部链接”中每个唯一商品占一行，商品链接可直接点击。\n"
        "2. 数据仅来自搜索结果卡片，页面未显示的字段保持空白。\n"
        "3. 价格和商品状态来自抓取时页面，可能随时变化。\n"
        "4. 遇到登录或安全验证时，软件只会暂停并等待人工处理。"
    )
    sheet["A13"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["A13"].fill = PatternFill("solid", fgColor="F7F7F7")
    sheet["A13"].font = Font(name="微软雅黑", color="5F5F5F")

    source_url = f"https://www.goofish.com/search?q={quote(config.keyword)}"
    sheet.merge_cells("A18:F18")
    sheet["A18"] = source_url
    sheet["A18"].hyperlink = source_url
    sheet["A18"].style = "Hyperlink"

    widths = [24, 34, 4, 24, 24, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(4, 11):
        if row != 7:
            sheet.row_dimensions[row].height = 27
    for row in range(13, 17):
        sheet.row_dimensions[row].height = 27


def _write_data_sheet(sheet, records: list[ProductRecord]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:R1")
    sheet["A1"] = "闲鱼商品链接"
    sheet["A1"].font = Font(name="微软雅黑", size=18, bold=True, color="1F2937")
    sheet["A1"].fill = PatternFill("solid", fgColor="F8D447")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 34

    sheet.merge_cells("A2:R2")
    sheet["A2"] = f"唯一商品：{len(records):,} 条｜每行均包含完整可点击链接"
    sheet["A2"].font = Font(name="微软雅黑", size=10, color="5F5F5F")
    sheet["A2"].fill = PatternFill("solid", fgColor="E9F7FF")
    sheet.row_dimensions[2].height = 23
    sheet.row_dimensions[3].height = 8

    header_fill = PatternFill("solid", fgColor="2B3745")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF")
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(4, column, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[4].height = 30

    thin_gray = Side(style="thin", color="E6E6E6")
    bottom_border = Border(bottom=thin_gray)
    for row_index, record in enumerate(records, start=5):
        values = [
            row_index - 4,
            record.keyword,
            record.item_id,
            record.title,
            record.price,
            record.original_price,
            record.region,
            record.condition,
            record.wants,
            record.reputation,
            record.publish_or_change,
            record.discount,
            record.first_page,
            record.appearances,
            ", ".join(str(page) for page in record.pages_seen),
            record.url,
            _as_datetime(record.captured_at),
            record.raw_text,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="微软雅黑", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=column in (4, 18))
            cell.border = bottom_border
        link_cell = sheet.cell(row_index, 16)
        link_cell.hyperlink = record.url
        link_cell.font = Font(name="微软雅黑", size=9, color="0563C1", underline="single")
        link_cell.alignment = Alignment(vertical="top")
        sheet.cell(row_index, 3).number_format = "@"
        sheet.cell(row_index, 5).number_format = '"¥"#,##0.00'
        sheet.cell(row_index, 6).number_format = '"¥"#,##0.00'
        sheet.cell(row_index, 17).number_format = "yyyy-mm-dd hh:mm:ss"
        sheet.row_dimensions[row_index].height = 36

    last_row = max(4, len(records) + 4)
    sheet.auto_filter.ref = f"A4:R{last_row}"
    sheet.freeze_panes = "C5"
    widths = [7, 18, 17, 48, 11, 11, 10, 14, 10, 16, 16, 16, 12, 10, 14, 48, 20, 70]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def export_workbook(
    *,
    config: CrawlConfig,
    records: list[ProductRecord],
    raw_records: int,
    searched_pages: int,
    stop_reason: str,
    started_at: datetime,
    finished_at: datetime,
    output_path: Path | None = None,
) -> Path:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    destination = output_path or _unique_output_path(config.output_dir, config.keyword, finished_at)
    destination = Path(destination).resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "汇总说明"
    data = workbook.create_sheet("全部链接")
    _style_summary(summary, config=config, row_count=len(records))
    _write_data_sheet(data, records)

    summary["E4"] = raw_records
    summary["E6"] = searched_pages
    summary["E7"] = stop_reason
    summary["E8"] = started_at
    summary["E9"] = finished_at
    summary["E8"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary["E9"].number_format = "yyyy-mm-dd hh:mm:ss"
    summary.freeze_panes = "A3"
    summary.sheet_properties.pageSetUpPr.fitToPage = True
    data.sheet_properties.pageSetUpPr.fitToPage = True
    data.page_setup.fitToWidth = 1
    data.page_setup.fitToHeight = 0
    data.print_title_rows = "1:4"
    workbook.active = 0
    workbook.save(destination)
    workbook.close()
    return destination


def export_monitor_workbook(
    task: MonitorTaskConfig,
    records: list[ProductRecord],
    output_dir: Path,
    *,
    output_path: Path | None = None,
) -> Path:
    """Exports the locally retained history for one monitor task on demand."""
    if any(not record.url.strip() for record in records):
        raise ValueError("监控商品链接不能为空")
    now = datetime.now()
    config = CrawlConfig(
        keyword=task.keyword,
        max_pages=task.pages,
        max_items=0,
        output_dir=output_dir,
        filters=task.filters,
    )
    destination = export_workbook(
        config=config,
        records=records,
        raw_records=len(records),
        searched_pages=task.pages,
        stop_reason=f"监控任务按需导出：{task.name}",
        started_at=now,
        finished_at=now,
        output_path=output_path,
    )
    workbook = load_workbook(destination)
    sheet = workbook["全部链接"]
    extra_headers = ("卖家ID", "聊天链接", "首次发现时间", "通知时间")
    for column, header in enumerate(extra_headers, start=19):
        cell = sheet.cell(4, column, header)
        cell.fill = PatternFill("solid", fgColor="2B3745")
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row, record in enumerate(records, start=5):
        sheet.cell(row, 19, record.seller_id)
        chat_cell = sheet.cell(row, 20, record.chat_url)
        if record.chat_url:
            chat_cell.hyperlink = record.chat_url
            chat_cell.font = Font(name="微软雅黑", size=9, color="0563C1", underline="single")
        sheet.cell(row, 21, _as_datetime(record.first_seen_at))
        sheet.cell(row, 22, _as_datetime(record.notified_at))
        sheet.cell(row, 21).number_format = "yyyy-mm-dd hh:mm:ss"
        sheet.cell(row, 22).number_format = "yyyy-mm-dd hh:mm:ss"
    for column, width in ((19, 20), (20, 52), (21, 20), (22, 20)):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = f"A4:V{max(4, len(records) + 4)}"
    workbook.save(destination)
    workbook.close()
    return destination
