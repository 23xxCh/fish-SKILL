from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from goofish_collector.exporter import export_monitor_workbook, export_workbook
from goofish_collector.models import CrawlConfig, ProductRecord, SearchFilters
from goofish_collector.monitor_models import MonitorTaskConfig


def make_records(count: int) -> list[ProductRecord]:
    return [
        ProductRecord(
            keyword="测试商品",
            item_id=str(10_000_000_000 + index),
            title=f"测试商品 {index}",
            price=float(index),
            url=f"https://www.goofish.com/item?id={10_000_000_000 + index}",
            first_page=index // 30 + 1,
            appearances=1,
            pages_seen=[index // 30 + 1],
            captured_at="2026-08-04 10:00:00",
            raw_text=f"测试商品 {index} ¥{index}",
        )
        for index in range(1, count + 1)
    ]


def test_exporter_creates_wps_compatible_workbook_with_all_links(tmp_path: Path) -> None:
    config = CrawlConfig(
        "测试商品",
        50,
        0,
        tmp_path,
        SearchFilters(
            min_price=100,
            max_price=800,
            region="广东",
            published_within="3天内",
            personal_only=True,
            inspection_only=True,
            free_shipping=True,
            brand_new=True,
        ),
    )
    records = make_records(1160)

    output = export_workbook(
        config=config,
        records=records,
        raw_records=1500,
        searched_pages=50,
        stop_reason="已到末页",
        started_at=datetime(2026, 8, 4, 10, 0, 0),
        finished_at=datetime(2026, 8, 4, 10, 30, 0),
    )

    workbook = load_workbook(output, read_only=False, data_only=False)
    assert workbook.sheetnames == ["汇总说明", "全部链接"]
    summary = workbook["汇总说明"]
    assert summary["B8"].value == "¥100.00 - ¥800.00"
    assert summary["B9"].value == "广东"
    assert summary["B10"].value == "3天内｜个人闲置、验货宝、包邮、全新"
    sheet = workbook["全部链接"]
    assert sheet.max_row == 1164
    assert sheet.max_column == 18
    ids = [sheet.cell(row, 3).value for row in range(5, sheet.max_row + 1)]
    links = [sheet.cell(row, 16).value for row in range(5, sheet.max_row + 1)]
    hyperlinks = [sheet.cell(row, 16).hyperlink for row in range(5, sheet.max_row + 1)]
    assert len(set(ids)) == 1160
    assert len(set(links)) == 1160
    assert all(links)
    assert all(link is not None and link.target == value for link, value in zip(hyperlinks, links))
    workbook.close()


def test_monitor_export_includes_real_chat_links(tmp_path: Path) -> None:
    task = MonitorTaskConfig(name="耳机新品", keyword="耳机")
    records = make_records(2)
    records[0].seller_id = "seller-1"
    records[0].chat_url = (
        "https://www.goofish.com/im?itemId=10000000001&peerUserId=seller-1"
    )
    records[0].first_seen_at = "2026-08-04 10:00:00"

    output = export_monitor_workbook(task, records, tmp_path)

    workbook = load_workbook(output)
    sheet = workbook["全部链接"]
    assert sheet.cell(4, 19).value == "卖家ID"
    assert sheet.cell(4, 20).value == "聊天链接"
    assert sheet.cell(5, 20).hyperlink.target == records[0].chat_url
    assert sheet.cell(6, 20).value in (None, "")
    assert all(sheet.cell(row, 16).hyperlink for row in (5, 6))
    workbook.close()
